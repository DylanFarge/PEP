import pandas as pd
import openpyxl as xl

class ExcelOutput:
    def __init__(self, groups, self_rate, lookup, capped_mark):
        self.capped_mark = capped_mark
        self.lookup = lookup        
        self.self_rate = self_rate
        self.df = pd.DataFrame(columns=["Ratings","Comments","Group"], index=groups["ID number"].tolist())

    def add_comment(self, rator, ratee, paragraph):
        comment = self.df.loc[rator]
        if type(comment["Comments"]) == dict:
            self.df.loc[rator]["Comments"][str(ratee)] = paragraph
        else:
            self.df.loc[rator]["Comments"] = {str(ratee): paragraph}

    def add_rating(self, rator, ratee, symbol):
        rating = self.df.loc[rator]
        if type(rating["Ratings"]) == dict:
            self.df.loc[rator]["Ratings"][str(ratee)] = symbol
        else:
            self.df.loc[rator]["Ratings"] = {str(ratee): symbol}
                                
    def _generate_excel_workbook(self, groups):

        def lookup_func(cell):
            string = "="
            for key, value in self.lookup.items():
                string += f'IF({cell} = "{key}", {value},'
            string += "0"
            for i in range(len(self.lookup)):
                string += ")"
            return string
    
        self.wb = xl.Workbook()

        for group in groups.groupby("Group"):
            st = self.wb.create_sheet(group[0])
            members = sorted(group[1]["ID number"].astype(int).tolist())
            
            for i, m in enumerate(members):
                st.cell(row=i+3, column=1, value=m)
                st.cell(row=2, column=(i*2)+2, value=m).alignment = xl.styles.Alignment(horizontal="center")
                st.merge_cells(start_row=2, start_column=(i*2)+2, end_row=2, end_column=(i*2)+3)

                m = str(m)
                if group[1][group[1]["ID number"] == m]["Flag"].values[0] == True:
                    print(f"Flagged {m}")
                    self.add_rating(m, m, [k for k,v in self.lookup.items() if v == 0][0])
                    for other in members:
                        if str(other) != m:
                            self.add_rating(m, other, [k for k,v in self.lookup.items() if v == 100][0])

                    # highlight whole row if flagged
                    for j in range(len(members)*3):
                        st.cell(row=i+3, column=j+1).fill = xl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                for j, n in enumerate(members):
                    n = str(n)
                    if m == n and not self.self_rate:
                        # Do something where we don't rate ourselves
                        st.cell(row=i+3, column=(j*2)+2, value="-").alignment = xl.styles.Alignment(horizontal="right")
                        st.cell(row=i+3, column=(j*2)+3, value="-").alignment = xl.styles.Alignment(horizontal="right")
                    
                    else:
                        cell = xl.utils.get_column_letter((j*2)+2) + str(i+3)
                        st.cell(row=i+3, column=(j*2)+2, value=self.df.loc[m]["Ratings"][n]).alignment = xl.styles.Alignment(horizontal="right")
                        st.cell(row=i+3, column=(j*2)+3, value=lookup_func(cell))

            # Heading
            st.cell(row=1, column=2, value="Group members BEING rated").alignment = xl.styles.Alignment(horizontal="center")
            st.merge_cells(start_row=1, start_column=2, end_row=1, end_column=(len(members)*2)+1)

            # Calculate student averages
            for k in range(len(members)):
                col = xl.utils.get_column_letter((k*2)+3)
                st.cell(row=i+6, column=(k*2)+2, value="=AVERAGE(" + col + "3:" + col + str(i+3) + ")").alignment = xl.styles.Alignment(horizontal="center")
                st.cell(row=i+5, column=(k*2)+2, value="Student Average").alignment = xl.styles.Alignment(horizontal="center")
                st.merge_cells(start_row=i+6, start_column=(k*2)+2, end_row=i+6, end_column=(k*2)+3)
                st.merge_cells(start_row=i+5, start_column=(k*2)+2, end_row=i+5, end_column=(k*2)+3)

            # Calculate team averages
            st.cell(row=i+5, column=(k*2)+5, value="Team Average").alignment = xl.styles.Alignment(horizontal="center")
            st.cell(row=i+6, column=(k*2)+5, value="=AVERAGE(A"+str(i+6)+":"+xl.utils.get_column_letter((k*2)+3)+""+str(i+6)+")").alignment = xl.styles.Alignment(horizontal="center")
            st.merge_cells(start_row=i+5, start_column=(k*2)+5, end_row=i+5, end_column=(k*2)+6)
            st.merge_cells(start_row=i+6, start_column=(k*2)+5, end_row=i+6, end_column=(k*2)+6)

            # Stating Capped Scaling Factor
            st.cell(row=2, column=(k*2)+8, value="Capped Scaling Factor").alignment = xl.styles.Alignment(horizontal="center")
            st.merge_cells(start_row=2, start_column=(k*2)+8, end_row=2, end_column=(k*2)+10)
            st.cell(row=3, column=(k*2)+8, value=float(self.capped_mark)).alignment = xl.styles.Alignment(horizontal="center")
            cap_cell = xl.utils.get_column_letter((k*2)+8) + str(3)
            st.merge_cells(start_row=3, start_column=(k*2)+8, end_row=3, end_column=(k*2)+10)

            # Calculate Scaling Factors
            st.cell(row=2, column=(k*2)+5, value="Scaling Factors").alignment = xl.styles.Alignment(horizontal="center")
            st.merge_cells(start_row=2, start_column=(k*2)+5, end_row=2, end_column=(k*2)+6)
            cell_team = xl.utils.get_column_letter((k*2)+5) + str(i+6)
            for m in range(len(members)):
                cell_student = xl.utils.get_column_letter((m*2)+2) + str(i+6)
                st.cell(row=(m+3), column=(k*2)+5, value="=MIN("+cell_student+"/"+cell_team+","+cap_cell+")").alignment = xl.styles.Alignment(horizontal="center")
                st.merge_cells(start_row=(m+3), start_column=(k*2)+5, end_row=(m+3), end_column=(k*2)+6)

            # Dump comments
            st.cell(row=i+6, column=(k*2) + 10, value="Comments").alignment = xl.styles.Alignment(horizontal="center")
            st.merge_cells(start_row=i+6, start_column=(k*2) + 10, end_row=i+6, end_column=(k*2) + 11)
            st.cell(row=i+7, column=(k*2) + 10, value="About").alignment = xl.styles.Alignment(horizontal="center")
            st.cell(row=i+7, column=(k*2) + 11, value="From").alignment = xl.styles.Alignment(horizontal="center")
            for about in range(len(members)):
                st.cell(row=i+8+(about * (len(members)+1)), column=(k*2) + 10, value=str(members[about])).alignment = xl.styles.Alignment(horizontal="center", vertical="center") 
                st.merge_cells(start_row=i+8+(about * (len(members)+1)), start_column=(k*2) + 10, end_row=i+8+((len(members)+1)*about)+(len(members)-1), end_column=(k*2) + 10)
                for frm in range(len(members)):
                    st.cell(row=i+8+((len(members)+1)*about) + frm, column=(k*2) + 11, value=str(members[frm])).alignment = xl.styles.Alignment(horizontal="center", vertical="center")
                    if group[1][group[1]["ID number"] == str(members[frm])]["Flag"].values[0] == True:
                        comment = "<Flagged Student>"        
                        st.cell(row=i+8+((len(members)+1)*about) + frm, column=(k*2) + 12, value=comment).fill = xl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    else:
                        comment = self.df.loc[str(members[frm])]["Comments"][str(members[about])]
                        st.cell(row=i+8+((len(members)+1)*about) + frm, column=(k*2) + 12, value=comment).alignment = xl.styles.Alignment(horizontal="left")

        # remove default sheet
        self.wb.remove(self.wb["Sheet"])
    
    def save(self, groups, dir):
        self._generate_excel_workbook(groups)
        self.wb.save(dir)